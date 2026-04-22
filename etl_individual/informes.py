"""
informes.py
Informe consolidado del ETL: informes/informe_etl.xlsx

Hojas:
  1_Resumen              — Dashboard ejecutivo con métricas clave
  2_Manifiestos          — Conteo raw vs cleaned por archivo + manifiestos perdidos
  3_Incoherencias        — Violaciones de reglas de negocio en el cleaned
  4_Diferencias          — Valores que cambiaron entre raw y cleaned
  5_Perfil_columnas      — Perfil por columna (tipos, nulos, rangos)
  6_Valores_categoricos  — Distribución de valores para columnas categóricas
  7_Linaje_columnas      — Todas las columnas raw y su destino (Conservada/Estandarizada/Fusionada/Transformada/Eliminada/Creada)
"""

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
from etl_individual.cleaning_individual import (
    RENAME_MAP,
    read_csv_with_header_detection,
    _strip_accents,
)

# ── Configuración ─────────────────────────────────────────────────────────────

DATA_FOLDER  = ROOT / "data_sheets"
CLEANED_DIR  = ROOT / "cleaned_data"
CLEANED_FILE = CLEANED_DIR / "individual_cleaned.csv"
OUTPUT_FILE  = ROOT / "informes/informe_etl.xlsx"

# Artefactos ETL generados por cleaning_individual.py
ETL_ARTIFACTS = [
    ("8_Filas_revision",             "filas_revision.csv"),
    ("9_Multi_remesas",              "multi_remesas.csv"),
    ("10_Estado_otros",              "estado_otros.csv"),
    ("11_Normalizacion_responsables","normalizacion_responsables.csv"),
]

NUMERIC_COLS   = ["valor_remesa", "flete_conductor", "valor_pagado", "anticipo"]
CITY_COLS      = ["origen", "destino"]
DATE_TEXT_COLS = ["fecha_despacho"]
PLAIN_COLS     = ["cliente", "placa", "conductor"]
TEXT_COLS      = CITY_COLS + DATE_TEXT_COLS + PLAIN_COLS

YEAR_MIN, YEAR_MAX  = 2015, 2030
ENUM_THRESHOLD      = 50
NOT_NULL_THRESHOLD  = 5.0

ID_COLS   = {"manifiesto", "consecutivo_semanal", "celular", "cedula_conductor",
             "factura_no", "factura_electronica", "remesas", "placa"}
DATE_COLS = {"periodo", "fecha_despacho", "fecha_cumplido", "fecha_pago", "fecha"}


# ── Helpers de normalización (comparación raw vs cleaned) ─────────────────────

def _to_float(v):
    """Convierte formatos monetarios colombianos a float."""
    if pd.isna(v):
        return None
    s = re.sub(r"[$\s*]", "", str(v).strip())
    if not s or s in ("-", "+"):
        return None
    if s.count(".") > 1:
        s = s.replace(".", "")
    if "," in s:
        if s.count(",") > 1:
            s = s.replace(",", "")
        elif re.search(r",\d{3}$", s):
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    elif re.match(r"^\d+\.\d{3}$", s):
        s = s.replace(".", "")
    try:
        return float(s)
    except ValueError:
        return None


def _norm_city(v) -> str:
    if pd.isna(v):
        return ""
    s = re.sub(r"\s*\([^)]*\)", "", str(v)).strip()
    s = _strip_accents(s)
    return re.sub(r"\s+", " ", s.upper())


def _norm_date(v) -> str:
    if pd.isna(v):
        return ""
    try:
        return str(pd.to_datetime(v, dayfirst=False).date())
    except Exception:
        return ""


def _norm_text(v) -> str:
    if pd.isna(v):
        return ""
    return re.sub(r"\s+", " ", str(v).strip().upper())


def _manif_key(val) -> str:
    if pd.isna(val):
        return ""
    s = re.sub(r"\s+", "", str(val).strip().upper())
    return re.sub(r"\.0$", "", s)


# ── Carga del cleaned ─────────────────────────────────────────────────────────

print("Cargando cleaned...")
cleaned = pd.read_csv(CLEANED_FILE, low_memory=False)

clean_manif_index = (
    cleaned.dropna(subset=["manifiesto"])
    .assign(_key=lambda d: d["manifiesto"].map(_manif_key))
    .set_index("_key")
)


# ── Comparación raw vs cleaned: manifiestos + diferencias ─────────────────────

rows_manifiestos = []
rows_diferencias = []
raw_col_registry: dict[str, list[str]] = {}   # col_normalizada → [archivos]

print("Procesando archivos raw...")
for path in sorted(DATA_FOLDER.glob("*.csv")):
    archivo = path.stem.strip()
    print(f"  {path.name}")

    try:
        raw = read_csv_with_header_detection(path)
    except Exception as e:
        print(f"    ERROR leyendo {path.name}: {e}")
        continue

    raw_cols = list(raw.columns)
    for _c in raw_cols:
        raw_col_registry.setdefault(_c, []).append(archivo)
    man_col  = next((c for c in raw_cols if "manifiesto" in c.lower()), None)

    if man_col is None:
        rows_manifiestos.append({
            "archivo": archivo, "total_raw": None, "total_clean": None,
            "diferencia": None, "anulados_raw": None, "perdidos": None,
            "manifiestos_perdidos": "", "nota": "Sin columna manifiesto en raw",
        })
        continue

    # Conteos
    valores_raw = raw[man_col].astype(str).str.strip().str.upper()
    anulados    = valores_raw.eq("ANULADO").sum()
    validos_raw = valores_raw[
        ~valores_raw.eq("ANULADO") &
        ~valores_raw.eq("NAN") &
        raw[man_col].notna() &
        valores_raw.str.match(r"^\d+(\.\d+)?$")
    ]

    clean_file  = cleaned[cleaned["archivo_origen"] == archivo]["manifiesto"]
    total_clean = int(clean_file.dropna().count())
    total_raw   = len(validos_raw)
    diferencia  = total_raw - total_clean

    # Manifiestos perdidos (integrado en la misma hoja)
    clean_keys = set(clean_file.dropna().map(_manif_key))
    perdidos_list = [v for v in validos_raw if _manif_key(v) and _manif_key(v) not in clean_keys]
    perdidos_str = ", ".join(perdidos_list[:15]) + (
        f"  … (+{len(perdidos_list) - 15} más)" if len(perdidos_list) > 15 else ""
    )

    rows_manifiestos.append({
        "archivo":       archivo,
        "total_raw":     total_raw,
        "total_clean":   total_clean,
        "diferencia":    diferencia,
        "anulados_raw":  int(anulados),
        "perdidos":      len(perdidos_list),
        "manifiestos_perdidos": perdidos_str,
        "nota": "OK" if diferencia == 0 else (
            f"{abs(diferencia)} calculados añadidos" if diferencia < 0
            else f"⚠ {diferencia} perdidos"
        ),
    })

    # Diferencias de valores
    raw["_key"] = raw[man_col].map(_manif_key)
    raw_indexed = raw[raw["_key"].str.len() > 0].set_index("_key")

    def raw_col_for(cleaned_name: str, _raw_cols=raw_cols):
        if cleaned_name in _raw_cols:
            return cleaned_name
        for raw_c, dest in RENAME_MAP.items():
            if dest == cleaned_name and raw_c in _raw_cols:
                return raw_c
        return None

    for clean_c in NUMERIC_COLS + TEXT_COLS:
        raw_c = raw_col_for(clean_c)
        if raw_c is None:
            continue

        for key in raw_indexed.index:
            if key not in clean_manif_index.index:
                continue

            raw_row   = raw_indexed.loc[key]
            clean_row = clean_manif_index.loc[key]
            if isinstance(raw_row,   pd.DataFrame): raw_row   = raw_row.iloc[0]
            if isinstance(clean_row, pd.DataFrame): clean_row = clean_row.iloc[0]

            raw_val   = raw_row[raw_c]     if raw_c   in raw_indexed.columns       else None
            clean_val = clean_row[clean_c] if clean_c in clean_manif_index.columns else None

            if clean_c in NUMERIC_COLS:
                rv, cv = _to_float(raw_val), _to_float(clean_val)
                if rv is None and cv is None: continue
                if rv == cv:                  continue
                diff = (cv - rv) if (cv is not None and rv is not None) else None
            else:
                if clean_c in CITY_COLS:
                    rv, cv = _norm_city(raw_val), _norm_city(clean_val)
                elif clean_c in DATE_TEXT_COLS:
                    rv, cv = _norm_date(raw_val), _norm_date(clean_val)
                else:
                    rv, cv = _norm_text(raw_val), _norm_text(clean_val)
                if rv == cv or (rv == "" and cv == ""):
                    continue
                diff = None

            rows_diferencias.append({
                "archivo":        archivo,
                "manifiesto":     key,
                "columna":        clean_c,
                "valor_raw":      raw_val,
                "valor_cleaned":  clean_val,
                "diferencia_num": diff,
            })


# ── Incoherencias (incluye antiguo "sin_remesa") ──────────────────────────────

print("Calculando incoherencias...")
rows_incoherencias = []


def _num(df, col):
    return pd.to_numeric(df[col], errors="coerce") if col in df.columns else pd.Series(pd.NA, index=df.index)


def _dt(df, col):
    return pd.to_datetime(df[col], errors="coerce") if col in df.columns else pd.Series(pd.NaT, index=df.index)


c_remesa   = _num(cleaned, "valor_remesa")
c_flete    = _num(cleaned, "flete_conductor")
c_pagado   = _num(cleaned, "valor_pagado")
c_anticipo = _num(cleaned, "anticipo")
c_dias_f   = _num(cleaned, "dias_para_facturar")
c_dias_c   = _num(cleaned, "dias_cumplido")


def _add_inco(mask, tipo, detalle_fn):
    for _, row in cleaned[mask].iterrows():
        rows_incoherencias.append({
            "archivo":    row.get("archivo_origen", ""),
            "manifiesto": row.get("manifiesto", ""),
            "tipo":       tipo,
            "detalle":    detalle_fn(row),
        })


_add_inco(c_flete.notna() & (c_flete > 0) & c_remesa.isna(),
    "Flete con remesa no digitada",
    lambda r: f"flete={r.get('flete_conductor')}  remesa=NaN")

_add_inco(c_flete.notna() & c_remesa.notna() & (c_flete > c_remesa) & (c_remesa > 0),
    "Flete > Remesa",
    lambda r: f"flete={r.get('flete_conductor')}  remesa={r.get('valor_remesa')}")

_add_inco(c_remesa.notna() & (c_remesa == 0),
    "valor_remesa = 0",
    lambda r: f"flete={r.get('flete_conductor')}")

for col, series in [("valor_remesa", c_remesa), ("flete_conductor", c_flete),
                    ("valor_pagado", c_pagado), ("anticipo", c_anticipo)]:
    _add_inco(series.notna() & (series < 0), f"{col} negativo",
        lambda r, c=col: f"{c}={r.get(c)}")

for col, series in [("flete_conductor", c_flete), ("valor_pagado", c_pagado), ("anticipo", c_anticipo)]:
    _add_inco(series.notna() & (series == 0), f"{col} = 0",
        lambda r, c=col: f"{c}={r.get(c)}")

_add_inco(c_dias_f.notna() & ((c_dias_f < -5) | (c_dias_f > 365)),
    "dias_para_facturar fuera de rango",
    lambda r: f"dias={r.get('dias_para_facturar')}")

_add_inco(c_dias_c.notna() & (c_dias_c < 0),
    "dias_cumplido negativo",
    lambda r: f"dias={r.get('dias_cumplido')}")

for date_col in ["fecha_despacho", "fecha_cumplido", "fecha_pago", "fecha"]:
    s = _dt(cleaned, date_col)
    _add_inco(s.notna() & ((s.dt.year < YEAR_MIN) | (s.dt.year > YEAR_MAX)),
        f"{date_col} fuera de rango ({YEAR_MIN}-{YEAR_MAX})",
        lambda r, c=date_col: f"{c}={r.get(c)}")


# ── Perfil de columnas ────────────────────────────────────────────────────────

print("Generando perfil de columnas...")


def _infer_type(col: str, series: pd.Series) -> str:
    if col in DATE_COLS: return "fecha"
    if col in ID_COLS:   return "id_texto"
    s = series.dropna()
    if s.empty:                                   return "vacio"
    if pd.api.types.is_bool_dtype(series):        return "booleano"
    if pd.api.types.is_integer_dtype(series):     return "entero"
    if pd.api.types.is_float_dtype(series):
        if s.apply(lambda x: float(x) == int(x)).all():
            return "entero"
        return "decimal"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "fecha"
    return "categorico" if s.nunique() <= ENUM_THRESHOLD else "texto_libre"


def _profile_column(col: str, series: pd.Series) -> dict:
    n_total   = len(series)
    n_nulos   = int(series.isna().sum())
    pct_nulos = round(n_nulos / n_total * 100, 2) if n_total else 0.0
    s         = series.dropna()
    tipo      = _infer_type(col, series)

    row = {
        "columna":           col,
        "tipo_inferido":     tipo,
        "dtype_pandas":      str(series.dtype),
        "n_total":           n_total,
        "n_nulos":           n_nulos,
        "pct_nulos":         pct_nulos,
        "not_null_sugerido": pct_nulos < NOT_NULL_THRESHOLD,
        "n_unicos":          int(s.nunique()),
        "enum_candidato":    tipo == "categorico",
        "min": None, "max": None, "media": None,
        "len_min": None, "len_max": None, "len_promedio": None,
        "valores_muestra":   "",
    }

    if s.empty:
        return row

    if tipo in ("entero", "decimal"):
        row["min"]   = round(float(s.min()), 4)
        row["max"]   = round(float(s.max()), 4)
        row["media"] = round(float(s.mean()), 4)
    elif tipo == "fecha":
        vals_sorted = sorted(v for v in s if pd.notna(v))
        if vals_sorted:
            row["min"] = str(vals_sorted[0])
            row["max"] = str(vals_sorted[-1])

    if tipo in ("categorico", "texto_libre", "id_texto"):
        lengths = s.astype(str).str.len()
        row["len_min"]      = int(lengths.min())
        row["len_max"]      = int(lengths.max())
        row["len_promedio"] = round(float(lengths.mean()), 1)

    unique_vals = sorted(str(v) for v in s.unique() if str(v).strip())
    if len(unique_vals) <= ENUM_THRESHOLD:
        row["valores_muestra"] = ", ".join(unique_vals)
    else:
        row["valores_muestra"] = ", ".join(unique_vals[:20]) + f"  … (+{len(unique_vals) - 20} más)"

    return row


# Parsear fechas conocidas antes del perfil
cleaned_for_profile = cleaned.copy()
for col in DATE_COLS:
    if col in cleaned_for_profile.columns:
        cleaned_for_profile[col] = pd.to_datetime(cleaned_for_profile[col], errors="coerce").dt.date

profile_rows = [_profile_column(col, cleaned_for_profile[col]) for col in cleaned_for_profile.columns]
profile_df   = pd.DataFrame(profile_rows)
tipo_map     = dict(zip(profile_df["columna"], profile_df["tipo_inferido"]))


# ── Valores categóricos ───────────────────────────────────────────────────────

print("Generando distribución de valores categóricos...")
n_total_rows = len(cleaned_for_profile)
enum_parts = []

for col in cleaned_for_profile.columns:
    if tipo_map.get(col) != "categorico":
        continue
    counts = (
        cleaned_for_profile[col].dropna()
        .astype(str)
        .value_counts()
        .reset_index()
    )
    counts.columns = ["valor", "frecuencia"]
    counts["columna"]       = col
    counts["pct_no_nulo"]   = (counts["frecuencia"] / cleaned_for_profile[col].notna().sum() * 100).round(2)
    counts["pct_del_total"] = (counts["frecuencia"] / n_total_rows * 100).round(2)
    enum_parts.append(counts[["columna", "valor", "frecuencia", "pct_no_nulo", "pct_del_total"]])

enum_df = pd.concat(enum_parts, ignore_index=True) if enum_parts else pd.DataFrame(
    columns=["columna", "valor", "frecuencia", "pct_no_nulo", "pct_del_total"]
)


# ── Linaje de columnas ────────────────────────────────────────────────────────

_TRANSFORMED_COLS_INFO: dict[str, str] = {
    "estado":                     "Categorizado con reglas ESTADO_RULES (PAGO A X DIAS, CONTRAENTREGA, PRONTO PAGO, URBANO, ANULADO…). Valores sin patrón → OTROS",
    "condicion_pago":             "Normalizada a categorías canónicas: CONTRAENTREGA, PAGO NORMAL, PRONTO PAGO, CONTINGENCIA 20-25 DH. Números sueltos → NULL",
    "entidad_financiera":         "Normalizada: método (TRANSF/CHEQUE) + banco (BANCOLOMBIA, DAVIVIENDA, BANCO DE BOGOTA). Personas y fechas coladas → NULL. Detalles extras → novedades",
    "estado_interno":             "Normalizada a valores canónicos del Enum (FACTURA RECIBIDA, NOVEDAD PENDIENTE, etc.)",
    "responsable":                "Lista negra de valores inusuales → NULL con nota en novedades. Valores cortos o con dígitos marcados como inusuales",
    "responsable_estado_interno": "Fuzzy-clustering (umbral 92 %) + correcciones explícitas de typos (LILANA→LILIANA, DAVIID→DAVID, CATHERN→CATHERIN, etc.)",
    "nombre_responsable":         "Fuzzy-clustering + correcciones explícitas (LILIANAOBREGON→LILIANA OBREGON, OPERAIVO 3→OPERATIVO 3, etc.)",
    "celular":                    "Validado: exactamente 10 dígitos. Inválidos → NULL con nota en novedades",
    "cedula_conductor":           "Validado: 6–12 dígitos. Inválidos → NULL con nota en novedades",
    "origen":                     "Nombre oficial estandarizado (ej. CALI→SANTIAGO DE CALI, BOGOTA→BOGOTÁ D.C.). Departamento extraído a columna separada. Paréntesis eliminados",
    "destino":                    "Nombre oficial estandarizado. Departamento extraído a columna separada. Paréntesis eliminados",
    "valor_remesa":               "Monetario colombiano → float ($ y puntos de miles eliminados). Multi-remesa (;): valores sumados",
    "flete_conductor":            "Monetario colombiano → float ($ y puntos de miles eliminados)",
    "anticipo":                   "Monetario colombiano → float",
    "valor_pagado":               "Monetario colombiano → float",
    "manifiesto":                 "ID limpiado: .0 eliminado, convertido a string entero",
    "consecutivo_semanal":        "ID limpiado: .0 eliminado, convertido a string entero",
    "fecha_despacho":             "Parseada a date (sin hora). Errores → NaT",
    "fecha_cumplido":             "Parseada a date (sin hora). Errores → NaT",
    "fecha_pago":                 "Parseada a date (sin hora). Errores → NaT",
    "fecha":                      "Parseada a date (sin hora). Errores → NaT",
    "dias_cumplido":              "Numérico limpiado: textos y fechas coladas → NULL. Valores fuera de rango (< -365 o > 3 650) → NULL",
    "agencia_despachadora":       "ANULADO (variantes) → 'ANULADO'. LETRAS+NÚMEROS → solo letras + nota. Lista negra (INFORMATIVO, RNDC) → NULL",
    "remesas":                    "Multi-remesa (';'): se consolida en una fila; valor_remesa acumulado; lista original conservada en este campo",
    "cliente":                    "Multi-remesa: clientes duplicados deduplicados dentro del mismo manifiesto",
    "mes_facturacion":            "Convertida a entero nullable (evita '4.0' de pandas)",
}

_CREATED_COLS_INFO: dict[str, str] = {
    "archivo_origen":       "Nombre del archivo CSV fuente sin extensión (trazabilidad)",
    "mes":                  "Mes en texto extraído del nombre del archivo (ej. 'ENERO')",
    "año":                  "Año numérico extraído del nombre del archivo",
    "periodo":              "Fecha del primer día del mes (YYYY-MM-01) derivada del nombre del archivo",
    "semana":               "Número de semana ('Semana 1'…) detectado de filas 'SEMANA X' en el CSV. 'N/A' si el archivo no usa semanas",
    "departamento_origen":  "Departamento extraído de origen: texto en paréntesis, palabras clave del nombre o tabla de fallback por ciudad conocida",
    "departamento_destino": "Departamento extraído de destino con la misma lógica que departamento_origen",
    "mes_facturacion":      "Mes numérico (1–12) de la fecha de factura (campo fecha.dt.month)",
    "dias_para_facturar":   "Días entre fecha_despacho y fecha de factura (puede ser negativo si hay error de fechas)",
    "novedades":            "Acumulado de notas de limpieza: entidad financiera ambigua, celular/cédula inválidos, estado especial, agencia inusual, etc.",
}

# Columnas que en DEAD_COLS se eliminan DESPUÉS de ser renombradas
_DEAD_COLS_FROM_RENAME: dict[str, str] = {
    "tiemp. lg cargue":    "tiempo_lg_cargue",
    "tiemp. lg descargue": "tiempo_lg_descargue",
}


def _build_linaje_df(col_registry: dict[str, list[str]]) -> pd.DataFrame:
    """Construye la tabla de linaje raw → cleaned para todas las columnas vistas."""
    from etl_individual.cleaning_individual import DROP_COLS, DEAD_COLS

    # Cuántos orígenes apuntan a cada destino (para detectar fusiones)
    dest_sources: dict[str, list[str]] = {}
    for src, dst in RENAME_MAP.items():
        dest_sources.setdefault(dst, []).append(src)
    fused_dests = {dst for dst, srcs in dest_sources.items() if len(srcs) > 1}

    rows = []

    for raw_col in sorted(col_registry.keys()):
        archivos = ", ".join(sorted(set(col_registry[raw_col])))
        col_key  = re.sub(r"\s+", " ", raw_col.strip()).lower()

        # 1. Columna de ruido / índice de Excel
        should_drop = (
            col_key in DROP_COLS
            or not col_key                                  # cadena vacía
            or col_key == "nan"                             # nombre NaN residual
            or re.fullmatch(r"\d+", col_key)
            or re.fullmatch(r"[a-z,]{1,2}", col_key)
            or bool(re.match(r"^unnamed:\s*\d+$", col_key))
        )
        if should_drop:
            rows.append({
                "columna_origen":  raw_col,
                "columna_destino": "—",
                "accion":          "Eliminada",
                "archivos_en":     archivos,
                "comentario":      "Columna de ruido, índice de Excel o nombre inválido → descartada",
            })
            continue

        # 2. Renombrada via RENAME_MAP
        if col_key in RENAME_MAP:
            dest    = RENAME_MAP[col_key]
            is_dead = dest in DEAD_COLS
            if is_dead:
                rows.append({
                    "columna_origen":  raw_col,
                    "columna_destino": "—",
                    "accion":          "Eliminada",
                    "archivos_en":     archivos,
                    "comentario":      f"Renombrada internamente a '{dest}', luego eliminada como columna muerta (demasiados nulos o cubierta por otra)",
                })
            else:
                is_fused = dest in fused_dests
                accion   = "Fusionada" if is_fused else "Estandarizada"
                base     = f"Renombrada a '{dest}'"
                if is_fused:
                    peers = ", ".join(f"'{s}'" for s in dest_sources[dest] if s != col_key)
                    base += f"; se fusiona con: {peers}"
                extra   = _TRANSFORMED_COLS_INFO.get(dest, "")
                comment = f"{base}. {extra}" if extra else base
                rows.append({
                    "columna_origen":  raw_col,
                    "columna_destino": dest,
                    "accion":          accion,
                    "archivos_en":     archivos,
                    "comentario":      comment,
                })
            continue

        # 3. Columna muerta que conserva su nombre (ej. agencia)
        if col_key in DEAD_COLS:
            extra_comment = (
                "Sus valores se propagan a 'agencia_despachadora' donde esté vacía antes de eliminarse"
                if col_key == "agencia"
                else "Demasiados nulos o cubierta por otra columna → eliminada"
            )
            rows.append({
                "columna_origen":  raw_col,
                "columna_destino": "—",
                "accion":          "Eliminada",
                "archivos_en":     archivos,
                "comentario":      extra_comment,
            })
            continue

        # 4. Columna generada por el ETL que también aparece en los raw
        #    (ej. semana creada por week_col, novedades pre-existente en algunos archivos)
        if col_key in _CREATED_COLS_INFO:
            rows.append({
                "columna_origen":  raw_col,
                "columna_destino": raw_col,
                "accion":          "Creada",
                "archivos_en":     archivos,
                "comentario":      _CREATED_COLS_INFO[col_key],
            })
            continue

        # 5. Columna que se conserva pero sus valores se transforman
        if col_key in _TRANSFORMED_COLS_INFO:
            rows.append({
                "columna_origen":  raw_col,
                "columna_destino": raw_col,
                "accion":          "Transformada",
                "archivos_en":     archivos,
                "comentario":      _TRANSFORMED_COLS_INFO[col_key],
            })
            continue

        # 6. Conservada sin cambios
        rows.append({
            "columna_origen":  raw_col,
            "columna_destino": raw_col,
            "accion":          "Conservada",
            "archivos_en":     archivos,
            "comentario":      "Sin cambios estructurales ni de valores",
        })

    # Columnas creadas por el ETL que NO aparecen en ningún raw
    for col, comment in _CREATED_COLS_INFO.items():
        if col in col_registry:
            continue   # ya se procesó arriba
        rows.append({
            "columna_origen":  "—",
            "columna_destino": col,
            "accion":          "Creada",
            "archivos_en":     "Todos los archivos",
            "comentario":      comment,
        })

    return pd.DataFrame(rows, columns=[
        "columna_origen", "columna_destino", "accion", "archivos_en", "comentario"
    ])


print("Construyendo linaje de columnas...")
df_linaje = _build_linaje_df(raw_col_registry)


# ── DataFrames de salida ──────────────────────────────────────────────────────

df_manifiestos   = pd.DataFrame(rows_manifiestos)
df_diferencias   = pd.DataFrame(rows_diferencias) if rows_diferencias else pd.DataFrame(
    columns=["archivo","manifiesto","columna","valor_raw","valor_cleaned","diferencia_num"])
df_incoherencias = pd.DataFrame(rows_incoherencias) if rows_incoherencias else pd.DataFrame(
    columns=["archivo","manifiesto","tipo","detalle"])


# ── Resumen (dashboard) ───────────────────────────────────────────────────────

print("Construyendo resumen...")

total_archivos  = len(df_manifiestos)
total_raw_sum   = int(df_manifiestos["total_raw"].fillna(0).sum())
total_clean_sum = int(df_manifiestos["total_clean"].fillna(0).sum())
total_perdidos  = int(df_manifiestos["perdidos"].fillna(0).sum())

inco_by_type = (
    df_incoherencias.groupby("tipo").size().reset_index(name="cantidad").sort_values("cantidad", ascending=False)
    if not df_incoherencias.empty else pd.DataFrame(columns=["tipo", "cantidad"])
)

diff_by_col = (
    df_diferencias.groupby("columna").size().reset_index(name="cantidad").sort_values("cantidad", ascending=False)
    if not df_diferencias.empty else pd.DataFrame(columns=["columna", "cantidad"])
)

cobertura_pct = round(total_clean_sum / total_raw_sum * 100, 2) if total_raw_sum else 0

resumen_rows = [
    ("— Cobertura ETL —", ""),
    ("Archivos procesados", total_archivos),
    ("Filas raw (con manifiesto válido)", total_raw_sum),
    ("Filas cleaned (con manifiesto)", total_clean_sum),
    ("Manifiestos perdidos", total_perdidos),
    ("Cobertura (%)", cobertura_pct),
    ("", ""),
    ("— Calidad de datos —", ""),
    ("Columnas totales", len(profile_df)),
    ("Columnas 100% vacías", int((profile_df["tipo_inferido"] == "vacio").sum())),
    (f"Candidatas a NOT NULL (<{NOT_NULL_THRESHOLD:.0f}% nulos)", int(profile_df["not_null_sugerido"].sum())),
    ("Candidatas a Enum/CHECK", int(profile_df["enum_candidato"].sum())),
    ("", ""),
    ("— Totales de problemas —", ""),
    ("Total incoherencias", len(df_incoherencias)),
    ("Total diferencias de valores (raw→cleaned)", len(df_diferencias)),
]
df_resumen = pd.DataFrame(resumen_rows, columns=["métrica", "valor"])


# ── Escribir Excel ────────────────────────────────────────────────────────────

OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
print(f"Escribiendo {OUTPUT_FILE}...")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df_resumen.to_excel(writer, sheet_name="1_Resumen", index=False)

    # Tablas auxiliares en la misma hoja de resumen
    row_offset = len(df_resumen) + 3
    if not inco_by_type.empty:
        inco_by_type.to_excel(writer, sheet_name="1_Resumen",
                              startrow=row_offset, index=False)
        row_offset += len(inco_by_type) + 3
    if not diff_by_col.empty:
        diff_by_col.to_excel(writer, sheet_name="1_Resumen",
                             startrow=row_offset, index=False)

    df_manifiestos.to_excel(writer,   sheet_name="2_Manifiestos",         index=False)
    df_incoherencias.to_excel(writer, sheet_name="3_Incoherencias",       index=False)
    df_diferencias.to_excel(writer,   sheet_name="4_Diferencias",         index=False)
    profile_df.to_excel(writer,       sheet_name="5_Perfil_columnas",     index=False)
    enum_df.to_excel(writer,          sheet_name="6_Valores_categoricos", index=False)
    df_linaje.to_excel(writer,        sheet_name="7_Linaje_columnas",     index=False)

    # Artefactos ETL (opcionales, generados por cleaning_individual.py)
    for sheet, filename in ETL_ARTIFACTS:
        path = CLEANED_DIR / filename
        if path.exists():
            pd.read_csv(path, low_memory=False).to_excel(writer, sheet_name=sheet, index=False)


# ── Formato ───────────────────────────────────────────────────────────────────

wb = load_workbook(OUTPUT_FILE)

HEADER_FILL  = PatternFill("solid", fgColor="2F5496")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, color="1F3864")
WARN_FILL    = PatternFill("solid", fgColor="FFEB9C")
ERROR_FILL   = PatternFill("solid", fgColor="FFC7CE")
OK_FILL      = PatternFill("solid", fgColor="C6EFCE")
CALC_FILL    = PatternFill("solid", fgColor="DDEBF7")


def style_header_row(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")


def autofit(ws, max_width=60):
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, max_width)


# — Hoja 1: Resumen —
ws = wb["1_Resumen"]
style_header_row(ws, 1)
for row in ws.iter_rows(min_row=2, max_row=len(df_resumen) + 1):
    val = str(row[0].value or "")
    if val.startswith("—"):
        for cell in row:
            cell.fill = SECTION_FILL
            cell.font = SECTION_FONT
# Estilar headers de tablas auxiliares
row_offset = len(df_resumen) + 4  # +3 por el gap + 1 por la fila header
if ws.cell(row=row_offset, column=1).value is not None:
    style_header_row(ws, row_offset)
if not inco_by_type.empty:
    row_offset2 = row_offset + len(inco_by_type) + 3
    if ws.cell(row=row_offset2, column=1).value is not None:
        style_header_row(ws, row_offset2)
autofit(ws)

# — Hoja 2: Manifiestos —
ws = wb["2_Manifiestos"]
style_header_row(ws)
diff_col_idx = next((i for i, c in enumerate(ws[1], 1) if c.value == "diferencia"), None)
if diff_col_idx:
    for row in ws.iter_rows(min_row=2):
        val = row[diff_col_idx - 1].value
        if val is None: continue
        row[diff_col_idx - 1].fill = CALC_FILL if val < 0 else (ERROR_FILL if val > 0 else OK_FILL)
autofit(ws)

# — Hoja 3: Incoherencias —
ws = wb["3_Incoherencias"]
style_header_row(ws)
for row in ws.iter_rows(min_row=2):
    tipo = str(row[2].value or "")
    fill = ERROR_FILL if ("negativo" in tipo or ">" in tipo or "fuera de rango" in tipo) else WARN_FILL
    for cell in row: cell.fill = fill
autofit(ws)

# — Hoja 4: Diferencias —
ws = wb["4_Diferencias"]
style_header_row(ws)
autofit(ws)

# — Hoja 5: Perfil columnas —
ws = wb["5_Perfil_columnas"]
style_header_row(ws)
autofit(ws)

# — Hoja 6: Valores categóricos —
ws = wb["6_Valores_categoricos"]
style_header_row(ws)
autofit(ws)

# — Hoja 7: Linaje de columnas —
_ACCION_FILLS = {
    "Estandarizada": PatternFill("solid", fgColor="C6EFCE"),   # verde claro
    "Fusionada":     PatternFill("solid", fgColor="FFEB9C"),   # amarillo
    "Transformada":  PatternFill("solid", fgColor="DDEBF7"),   # azul claro
    "Eliminada":     PatternFill("solid", fgColor="FFC7CE"),   # rojo claro
    "Creada":        PatternFill("solid", fgColor="E2EFDA"),   # verde más suave
    "Conservada":    None,
}
ws = wb["7_Linaje_columnas"]
style_header_row(ws)
accion_col_idx = next(
    (i for i, c in enumerate(ws[1], 1) if str(c.value or "").lower() == "accion"), None
)
if accion_col_idx:
    for row in ws.iter_rows(min_row=2):
        accion_val = str(row[accion_col_idx - 1].value or "")
        fill = _ACCION_FILLS.get(accion_val)
        if fill:
            for cell in row:
                cell.fill = fill
autofit(ws)

# — Hojas de artefactos ETL —
for sheet, _ in ETL_ARTIFACTS:
    if sheet in wb.sheetnames:
        ws = wb[sheet]
        style_header_row(ws)
        autofit(ws)

wb.save(OUTPUT_FILE)


# ── Resumen consola ──────────────────────────────────────────────────────────

print(f"\nInforme guardado en: {OUTPUT_FILE}")
print(f"  Archivos procesados:     {total_archivos}")
print(f"  Filas raw / cleaned:     {total_raw_sum} / {total_clean_sum}  (cobertura {cobertura_pct}%)")
print(f"  Manifiestos perdidos:    {total_perdidos}")
print(f"  Diferencias de valores:  {len(df_diferencias)}")
print(f"  Incoherencias:           {len(df_incoherencias)}")
print(f"  Columnas analizadas:     {len(profile_df)}")
