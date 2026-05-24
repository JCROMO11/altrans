"""
Compara el CSV cleaned (o la DB actual) contra un Excel fuente
(Lista_Manifiestos_*.xlsx), aplicando el mismo mapeo de columnas y normalización
que el panel de carga del dashboard (buildPayload + normalizeVal).

Genera un Excel con 4 hojas en tests/reportes/:
  1. Resumen            — conteos globales
  2. Con Cambios        — manifiesto, campo, valor_anterior, valor_nuevo
  3. Nuevos             — manifiestos que están en el Excel pero NO en la fuente
  4. Sin Cambios        — manifiestos idénticos en ambas fuentes

Uso:
  # Comparar CSV cleaned vs Excel (pre-upload: predice qué va a cambiar)
  python3 tests/generar_excel_comparacion.py

  # Comparar DB actual vs Excel (post-upload: verifica que los cambios se aplicaron)
  python3 tests/generar_excel_comparacion.py --desde-db

  # Archivos personalizados
  python3 tests/generar_excel_comparacion.py --csv otro.csv --excel otro.xlsx
"""

import argparse
import csv
import math
import os
import re
import unicodedata
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constantes (coinciden con CargaPage.jsx y normalize.js)
# ---------------------------------------------------------------------------

DB_FIELDS = [
    "fecha_despacho", "origen", "destino", "cliente", "conductor",
    "cedula_conductor", "celular", "placa", "tipo_vehiculo", "propietario",
    "agencia_despachadora", "nombre_responsable",
    "valor_remesa", "flete_conductor", "anticipo", "remesas",
]

FIELD_LABELS = {
    "fecha_despacho": "Fecha despacho", "origen": "Origen", "destino": "Destino",
    "cliente": "Cliente", "conductor": "Conductor", "cedula_conductor": "Cédula",
    "celular": "Celular", "placa": "Placa", "tipo_vehiculo": "Remolque",
    "propietario": "Propietario", "agencia_despachadora": "Agencia Despachadora",
    "nombre_responsable": "Responsable",
    "valor_remesa": "Valor remesa", "flete_conductor": "Flete",
    "anticipo": "Anticipo", "remesas": "Remesas",
}

NUMERIC_FIELDS = {"valor_remesa", "flete_conductor", "anticipo"}

# ---------------------------------------------------------------------------
# Mapeo de columnas Excel → campo DB (réplica de buildPayload en excel-upload.js)
# ---------------------------------------------------------------------------

EXCEL_TO_DB = {
    "MANIFIESTO":         "manifiesto",
    "FECHA EMISIÓN":      "fecha_despacho",
    "AGENCIA":            "agencia_despachadora",
    "ORIGEN":             "origen",
    "DESTINO":            "destino",
    "PLACA":              "placa",
    "REMOLQUE":           "tipo_vehiculo",
    "REMESAS":            "remesas",
    "VALORES REMESAS":    "valor_remesa",
    "CONDUCTOR":          "conductor",
    "DOC. CONDUCTOR":     "cedula_conductor",
    "TEL. CONDUCTOR":     "celular",
    "GENERADORES":        "cliente",
    "FLETE":              "flete_conductor",
    "ANTICIPO":           "anticipo",
    "CREADO POR":         "nombre_responsable",
    "POSEEDOR":           "propietario",      # POSEEDOR tiene prioridad sobre PROPIETARIO
    "PROPIETARIO":        "propietario",
}

# Mapa de abreviaturas de departamento (geography.js DEPT_ABBREV)
DEPT_ABBREV = {
    'Anti': 'Antioquia', 'Atla': 'Atlantico', 'Bogo': 'Bogota D.C.',
    'Boli': 'Bolivar', 'Boya': 'Boyaca', 'Cald': 'Caldas',
    'Casa': 'Casanare', 'Cauc': 'Cauca', 'Cesa': 'Cesar',
    'Cord': 'Cordoba', 'Cund': 'Cundinamarca', 'Huil': 'Huila',
    'La G': 'La Guajira', 'Magd': 'Magdalena', 'Meta': 'Meta',
    'Nari': 'Nariño', 'Nort': 'Norte de Santander', 'Quin': 'Quindio',
    'Risa': 'Risaralda', 'Sant': 'Santander', 'Toli': 'Tolima',
    'Vall': 'Valle del Cauca', 'Arau': 'Arauca', 'Caqu': 'Caqueta',
    'Guav': 'Guaviare', 'Putu': 'Putumayo', 'Sucr': 'Sucre',
}

RESPONSABLE_FIXES = {
    'OPERATIVO3': 'OPERATIVO 3',
    'OPERAIVO 3': 'OPERATIVO 3',
    'LILIANAOBREGON': 'LILIANA OBREGON',
    'VANESA': 'VANESSA',
}

# ---------------------------------------------------------------------------
# Réplicas en Python de las funciones del frontend
# ---------------------------------------------------------------------------

def _strip_accents(s):
    """Elimina acentos manteniendo ñ/Ñ, igual que removeAccents() en normalize.js."""
    if not s:
        return s
    s = s.replace('ñ', '\x00').replace('Ñ', '\x01')
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if not unicodedata.category(c).startswith('M'))
    s = s.replace('\x00', 'ñ').replace('\x01', 'Ñ')
    return s


def parse_ciudad(v):
    """Extrae la ciudad de un valor 'CIUDAD(Depto)' como parseCiudad() en geography.js."""
    if not v:
        return None
    s = _strip_accents(str(v).strip())
    m = re.match(r'^(.*?)\s*\(([^)]+)\)\s*$', s)
    if m:
        return m.group(1).strip()
    return s


def parse_fecha(v):
    """Convierte fecha Excel a 'YYYY-MM-DD', réplica de parseFecha()."""
    if v is None or v == '':
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d') if hasattr(v, 'strftime') else str(v)[:10]
    if isinstance(v, (int, float)):
        from datetime import timedelta
        d = datetime(1899, 12, 30) + timedelta(days=int(v))
        return d.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s:
        return None
    return s[:10]


def to_num(v):
    """Convierte valores numéricos Colombian/Excel, réplica de toNum()."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        if math.isnan(v):
            return None
        return v if v != 0 else None
    s = str(v).strip()
    if not s:
        return None
    if ';' in s:
        suma = sum((to_num(part.strip()) or 0) for part in s.split(';'))
        return suma or None
    con_coma = ',' in s
    con_punto = '.' in s
    if con_coma and con_punto:
        last_comma = s.rfind(',')
        last_punto = s.rfind('.')
        if last_comma > last_punto:
            limpio = s.replace('.', '').replace(',', '.')
        else:
            limpio = s.replace(',', '')
    elif con_coma:
        partes = s.split(',')
        if len(partes) > 1 and len(partes[-1]) == 2:
            limpio = s.rsplit(',', 1)[0].replace(',', '') + '.' + s.rsplit(',', 1)[1]
        else:
            limpio = s.replace(',', '')
    elif con_punto:
        partes = s.split('.')
        if len(partes) > 1 and len(partes[-1]) == 2:
            limpio = s
        else:
            limpio = s.replace('.', '')
    else:
        limpio = s
    try:
        n = float(limpio)
    except (ValueError, TypeError):
        return None
    if math.isnan(n) or n == 0:
        return None
    return n


def normalize_responsable(s):
    """Normaliza nombre de responsable, réplica de normalizeResponsable()."""
    if s is None or s == '':
        return None
    up = str(s).strip().upper()
    if not up:
        return None
    return RESPONSABLE_FIXES.get(up, up)


# ---------------------------------------------------------------------------
# normalizeVal — réplica exacta de normalize.js
# ---------------------------------------------------------------------------

def normalize_val(v, field):
    """Normaliza un valor para comparación, igual que normalizeVal() en el frontend."""
    if v is None or v == '':
        return None
    if field in NUMERIC_FIELDS:
        try:
            n = float(v)
        except (ValueError, TypeError):
            return None
        if math.isnan(n) or n == 0:
            return None
        return str(round(n))
    return str(v).strip()


# ---------------------------------------------------------------------------
# Carga de archivos
# ---------------------------------------------------------------------------

def cargar_csv(path):
    """Lee un CSV y devuelve lista de diccionarios con claves originales."""
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        return list(reader)


def cargar_excel_fuente(path):
    """
    Lee la primera hoja de un .xlsx fuente (formato Lista_Manifiestos)
    y devuelve una lista de diccionarios con las columnas raw del Excel.
    """
    import pandas as pd
    xl = pd.ExcelFile(path, engine='calamine')
    ws = xl.sheet_names[0]
    df = xl.parse(ws)
    rows = []
    for _, row in df.iterrows():
        d = {}
        for col in df.columns:
            val = row[col]
            if isinstance(val, float) and math.isnan(val):
                d[col] = None
            else:
                d[col] = val
        rows.append(d)
    return rows


def cargar_desde_db(manifiestos_lista=None):
    """
    Exporta los campos relevantes de manifiestos_flat directamente desde Supabase.
    Si manifiestos_lista es una lista de enteros, solo trae esos manifiestos
    (más eficiente que traer toda la tabla).
    Retorna lista de dicts con las mismas claves que un CSV row.
    """
    from dotenv import load_dotenv
    import psycopg2
    import psycopg2.extras

    load_dotenv()
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL no encontrado en .env")

    cols = ["manifiesto"] + DB_FIELDS
    cols_sql = ", ".join(cols)

    conn = psycopg2.connect(db_url)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if manifiestos_lista:
        placeholders = ",".join(["%s"] * len(manifiestos_lista))
        cur.execute(
            f"SELECT {cols_sql} FROM public.manifiestos_flat WHERE manifiesto IN ({placeholders})",
            manifiestos_lista,
        )
    else:
        cur.execute(f"SELECT {cols_sql} FROM public.manifiestos_flat")

    rows = []
    for db_row in cur.fetchall():
        r = {}
        for col in cols:
            val = db_row[col]
            if val is None:
                r[col] = ""
            elif hasattr(val, 'isoformat'):
                r[col] = val.isoformat()[:10]
            else:
                r[col] = str(val)
        rows.append(r)

    cur.close()
    conn.close()
    return rows


# ---------------------------------------------------------------------------
# Conversión de fila Excel → formato comparable (mismo mapeo que buildPayload)
# ---------------------------------------------------------------------------

def excel_row_to_comparable(excel_row):
    """
    Convierte una fila del Excel fuente a un diccionario con los 16 DB_FIELDS,
    aplicando el mismo mapeo y normalización que buildPayload() en el frontend.
    """
    result = {}

    raw_m = excel_row.get('MANIFIESTO')
    try:
        manifiesto = int(float(str(raw_m).strip()))
    except (ValueError, TypeError):
        return None
    result['manifiesto'] = manifiesto

    result['fecha_despacho'] = parse_fecha(excel_row.get('FECHA EMISIÓN'))

    result['origen'] = parse_ciudad(excel_row.get('ORIGEN'))
    result['destino'] = parse_ciudad(excel_row.get('DESTINO'))

    gen = excel_row.get('GENERADORES')
    if gen is not None and str(gen).strip():
        result['cliente'] = str(gen).split(';')[0].strip()
    else:
        result['cliente'] = None

    result['conductor'] = _trim_or_none(excel_row.get('CONDUCTOR'))
    result['cedula_conductor'] = _trim_or_none(excel_row.get('DOC. CONDUCTOR'))
    result['celular'] = _trim_or_none(excel_row.get('TEL. CONDUCTOR'))
    result['placa'] = _trim_or_none(excel_row.get('PLACA'))
    result['tipo_vehiculo'] = _trim_or_none(excel_row.get('REMOLQUE'))

    poseedor = excel_row.get('POSEEDOR')
    propietario = excel_row.get('PROPIETARIO')
    result['propietario'] = _trim_or_none(poseedor) or _trim_or_none(propietario)

    result['agencia_despachadora'] = _trim_or_none(excel_row.get('AGENCIA'))
    result['nombre_responsable'] = normalize_responsable(excel_row.get('CREADO POR'))

    result['valor_remesa'] = to_num(excel_row.get('VALORES REMESAS'))
    result['flete_conductor'] = to_num(excel_row.get('FLETE'))
    result['anticipo'] = to_num(excel_row.get('ANTICIPO'))
    result['remesas'] = _trim_or_none(excel_row.get('REMESAS'))

    return result


def _trim_or_none(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Comparación
# ---------------------------------------------------------------------------

def comparar(csv_rows, excel_rows):
    """
    Compara cada fila del Excel (mapeada/normalizada) contra el CSV/DB indexado
    por manifiesto, usando los 16 DB_FIELDS con normalize_val().
    """
    csv_map = {}
    for r in csv_rows:
        try:
            m = int(r.get('manifiesto', 0))
        except (ValueError, TypeError):
            continue
        csv_map[m] = r

    excel_comparables = []
    errores = []
    for row in excel_rows:
        comp = excel_row_to_comparable(row)
        if comp is None:
            errores.append(row)
        else:
            excel_comparables.append(comp)

    nuevos = []
    sin_cambios = []
    con_cambios = []

    for exc in excel_comparables:
        m = exc['manifiesto']
        csv_row = csv_map.get(m)

        if csv_row is None:
            nuevos.append(exc)
            continue

        diffs = []
        for f in DB_FIELDS:
            val_csv = normalize_val(csv_row.get(f), f)
            val_exc = normalize_val(exc.get(f), f)
            if val_csv != val_exc:
                diffs.append({
                    "field": f,
                    "label": FIELD_LABELS.get(f, f),
                    "val_anterior": csv_row.get(f),
                    "val_nuevo": exc.get(f),
                    "val_anterior_norm": val_csv,
                    "val_nuevo_norm": val_exc,
                })

        if not diffs:
            sin_cambios.append(exc)
        else:
            con_cambios.append({"row": exc, "diffs": diffs})

    return {
        "nuevos": nuevos,
        "sin_cambios": sin_cambios,
        "con_cambios": con_cambios,
        "errores": errores,
    }


# ---------------------------------------------------------------------------
# Generar Excel de salida
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
SUMMARY_LABEL_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_VALUE_FONT = Font(name="Calibri", size=11)
NEW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CHANGED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, max_width=55):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.value:
                for line in str(cell.value).split('\n'):
                    lengths.append(len(line))
        best = min(max(lengths, default=10) + 3, max_width)
        ws.column_dimensions[col_letter].width = best


def _style_data_rows(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def generar_excel(resultado, output_path, fuente_label="CSV"):
    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"

    n_nuevos = len(resultado["nuevos"])
    n_cambios = len(resultado["con_cambios"])
    n_iguales = len(resultado["sin_cambios"])
    n_errores = len(resultado.get("errores", []))
    total = n_nuevos + n_cambios + n_iguales

    ws1.merge_cells("A1:B1")
    ws1["A1"] = "Comparación de Manifiestos — Resumen"
    ws1["A1"].font = TITLE_FONT

    ws1.merge_cells("A2:B2")
    ws1["A2"] = f"Fuente de comparación: {fuente_label}"
    ws1["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")

    resumen = [
        ("Total manifiestos en archivo Excel cargado", total),
        ("", ""),
        ("Sin cambios (idénticos en ambas fuentes)", n_iguales),
        ("Con cambios (diferencias detectadas)", n_cambios),
        ("Nuevos (están en el Excel pero NO en la fuente)", n_nuevos),
    ]
    if n_errores > 0:
        resumen.append(("", ""))
        resumen.append(("Filas con error (sin manifiesto válido)", n_errores))

    for i, (label, val) in enumerate(resumen, start=4):
        ws1.cell(row=i, column=1, value=label).font = SUMMARY_LABEL_FONT
        c = ws1.cell(row=i, column=2, value=val)
        c.font = SUMMARY_VALUE_FONT
        if "Sin cambios" in label:
            c.font = Font(name="Calibri", size=11, color="006100")
        elif "Con cambios" in label:
            c.font = Font(name="Calibri", size=11, color="BF8F00")
        elif "Nuevos" in label:
            c.font = Font(name="Calibri", size=11, color="0066CC")

    ws1.column_dimensions["A"].width = 50
    ws1.column_dimensions["B"].width = 15

    # ── Hoja 2: Con Cambios ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Con Cambios")

    headers2 = ["Manifiesto", "Campo", f"Valor Anterior ({fuente_label})", "Valor Nuevo (Excel)"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, len(headers2))

    row = 2
    for item in resultado["con_cambios"]:
        m = item["row"].get("manifiesto", "")
        for d in item["diffs"]:
            ws2.cell(row=row, column=1, value=m)
            ws2.cell(row=row, column=2, value=d["label"])
            ws2.cell(row=row, column=3, value=_fmt(d["val_anterior"]))
            ws2.cell(row=row, column=4, value=_fmt(d["val_nuevo"]))
            for c in range(1, 5):
                ws2.cell(row=row, column=c).fill = CHANGED_FILL
            row += 1

    _style_data_rows(ws2, 2, row - 1, len(headers2))
    _auto_width(ws2)

    # ── Hoja 3: Nuevos ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Nuevos")

    headers3 = ["Manifiesto"] + [FIELD_LABELS.get(f, f) for f in DB_FIELDS]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    _style_header(ws3, len(headers3))

    for i, p in enumerate(resultado["nuevos"], start=2):
        ws3.cell(row=i, column=1, value=p.get("manifiesto", ""))
        for j, f in enumerate(DB_FIELDS, start=2):
            ws3.cell(row=i, column=j, value=_fmt(p.get(f)))
        for c in range(1, len(headers3) + 1):
            ws3.cell(row=i, column=c).fill = NEW_FILL

    _style_data_rows(ws3, 2, len(resultado["nuevos"]) + 1, len(headers3))
    _auto_width(ws3)

    # ── Hoja 4: Sin Cambios ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Sin Cambios")

    headers4 = ["Manifiesto"] + [FIELD_LABELS.get(f, f) for f in DB_FIELDS]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=c, value=h)
    _style_header(ws4, len(headers4))

    for i, p in enumerate(resultado["sin_cambios"], start=2):
        ws4.cell(row=i, column=1, value=p.get("manifiesto", ""))
        for j, f in enumerate(DB_FIELDS, start=2):
            ws4.cell(row=i, column=j, value=_fmt(p.get(f)))

    _style_data_rows(ws4, 2, len(resultado["sin_cambios"]) + 1, len(headers4))
    _auto_width(ws4)

    wb.save(output_path)
    return output_path


def _fmt(v):
    if v is None:
        return "(vacío)"
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Compara CSV cleaned (o DB actual) vs Excel fuente de manifiestos"
    )
    parser.add_argument("--csv", type=str, default=None,
                        help="CSV cleaned (default: cleaned_data/individual_cleaned.csv)")
    parser.add_argument("--excel", type=str, default=None,
                        help="Excel fuente (default: data/Lista_Manifiestos_08_05_2026.xlsx)")
    parser.add_argument("--output", type=str, default=None,
                        help="Ruta del Excel de salida (default: tests/reportes/comparacion.xlsx)")
    parser.add_argument("--desde-db", action="store_true",
                        help="Leer estado actual desde Supabase en vez del CSV "
                             "(usar DESPUÉS del excel-upload para verificar que los cambios se aplicaron)")
    args = parser.parse_args()

    reportes_dir = os.path.join(os.path.dirname(__file__), "reportes")
    os.makedirs(reportes_dir, exist_ok=True)

    if args.desde_db:
        suffix = "post_upload"
        fuente_label = "DB actual"
    else:
        suffix = "pre_upload"
        fuente_label = "CSV cleaned"

    output_path = args.output or os.path.join(reportes_dir, f"comparacion_{suffix}.xlsx")

    excel_path = args.excel or "data/Lista_Manifiestos_08_05_2026.xlsx"

    print(f"Cargando Excel fuente: {excel_path}")
    excel_rows = cargar_excel_fuente(excel_path)
    print(f"  {len(excel_rows)} filas en Excel")

    if args.desde_db:
        # Extrae los manifiestos del Excel para hacer una query dirigida
        excel_manifiestos = []
        for row in excel_rows:
            raw_m = row.get('MANIFIESTO')
            try:
                excel_manifiestos.append(int(float(str(raw_m).strip())))
            except (ValueError, TypeError):
                pass
        print(f"Consultando DB para {len(excel_manifiestos)} manifiestos...")
        csv_rows = cargar_desde_db(excel_manifiestos)
        print(f"  {len(csv_rows)} manifiestos encontrados en DB")
    else:
        csv_path = args.csv or "cleaned_data/individual_cleaned.csv"
        print(f"Cargando CSV cleaned: {csv_path}")
        csv_rows = cargar_csv(csv_path)
        print(f"  {len(csv_rows)} filas en CSV")

    resultado = comparar(csv_rows, excel_rows)

    n_nuevos = len(resultado["nuevos"])
    n_cambios = len(resultado["con_cambios"])
    n_iguales = len(resultado["sin_cambios"])
    n_errores = len(resultado.get("errores", []))
    total_diffs = sum(len(item["diffs"]) for item in resultado["con_cambios"])

    print(f"\nResultados ({fuente_label} vs Excel):")
    print(f"  Sin cambios : {n_iguales}")
    print(f"  Con cambios : {n_cambios} manifiestos ({total_diffs} campos diferentes)")
    print(f"  Nuevos      : {n_nuevos}")
    if n_errores:
        print(f"  Errores     : {n_errores} filas sin manifiesto válido")

    generar_excel(resultado, output_path, fuente_label=fuente_label)
    print(f"\nReporte generado en: {output_path}")


if __name__ == "__main__":
    main()
