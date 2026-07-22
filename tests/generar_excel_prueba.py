"""
Genera Excels de prueba para verificar el comportamiento del panel de carga
del dashboard ante casos límite (filas inválidas y archivo vacío).

Uso:
  # Excel con 10 filas (6 válidas + 4 inválidas)
  python3 tests/generar_excel_prueba.py --invalidas

  # Excel vacío (solo cabeceras, sin datos)
  python3 tests/generar_excel_prueba.py --vacio

  # Ambos de una vez
  python3 tests/generar_excel_prueba.py --invalidas --vacio

Los archivos se guardan en tests/reportes/.
"""

import argparse
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Columnas en el mismo orden que el Excel fuente real (Lista_Manifiestos)
HEADERS = [
    "MANIFIESTO", "FECHA EMISIÓN", "AGENCIA", "ORIGEN", "DESTINO",
    "PLACA", "TIPO DE VEHICULO", "REMESAS", "VALORES REMESAS",
    "CONDUCTOR", "DOC. CONDUCTOR", "TEL. CONDUCTOR",
    "GENERADORES", "FLETE", "ANTICIPO", "CREADO POR",
    "POSEEDOR", "PROPIETARIO",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
INVALID_FILL = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _write_headers(ws):
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _write_row(ws, row_num, values, invalida=False):
    fill = INVALID_FILL if invalida else None
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def _auto_width(ws):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = [len(str(cell.value)) for cell in col_cells if cell.value]
        ws.column_dimensions[col_letter].width = min(max(lengths, default=10) + 3, 40)


def _fila(manifiesto, fecha, agencia, origen, destino, placa, remolque,
          remesas, valor_remesa, conductor, cedula, celular,
          cliente, flete, anticipo, responsable, poseedor="", propietario=""):
    return [
        manifiesto, fecha, agencia, origen, destino,
        placa, remolque, remesas, valor_remesa,
        conductor, cedula, celular,
        cliente, flete, anticipo, responsable,
        poseedor, propietario,
    ]


# ---------------------------------------------------------------------------
# Filas de prueba
# ---------------------------------------------------------------------------

# Manifiestos con números muy altos para que sean "nuevos" (no existen en DB)
FILAS_VALIDAS = [
    _fila(999001, "2026-01-15", "AGENCIA NORTE", "BOGOTA(Cund)", "CUCUTA(Nort)",
          "ABC123", "SENCILLO", "REM-5001", 1500000,
          "JUAN PEREZ", "12345678", "3101234567",
          "CLIENTE ALFA", 350000, 50000, "ANA LOPEZ"),
    _fila(999002, "2026-01-16", "AGENCIA SUR", "MEDELLIN(Anti)", "BARRANQUILLA(Atla)",
          "DEF456", "DOBLE TROQUE", "REM-5002", 2200000,
          "CARLOS GOMEZ", "87654321", "3209876543",
          "COMERCIAL BETA", 480000, 0, "MARIA GARCIA"),
    _fila(999003, "2026-02-03", "AGENCIA NORTE", "CALI(Vall)", "SANTA MARTA(Magd)",
          "GHI789", "SENCILLO", "REM-5003;REM-5004", "800000;700000",
          "ANDRES RODRIGUEZ", "11223344", "3154445566",
          "INDUSTRIAS GAMMA", 310000, 100000, "ANA LOPEZ"),
    _fila(999004, "2026-02-10", "AGENCIA SUR", "BOGOTA(Cund)", "PEREIRA(Risa)",
          "JKL012", "", "REM-5005", 950000,
          "LUIS MARTINEZ", "55667788", "3007778899",
          "CLIENTE ALFA", 270000, 0, "OPERATIVO 3"),
    _fila(999005, "2026-03-01", "AGENCIA NORTE", "BARRANQUILLA(Atla)", "MANIZALES(Cald)",
          "MNO345", "MINI MULA", "REM-5006", 3100000,
          "PEDRO SANCHEZ", "99001122", "3112223344",
          "COMERCIAL BETA", 560000, 150000, "LILIANA OBREGON"),
    _fila(999006, "2026-03-20", "AGENCIA SUR", "CUCUTA(Nort)", "BOGOTA(Cund)",
          "PQR678", "SENCILLO", "REM-5007", 1750000,
          "JORGE CASTRO", "44332211", "3188887766",
          "INDUSTRIAS GAMMA", 400000, 80000, "VANESSA"),
]

# Filas inválidas — distintos tipos de error
FILAS_INVALIDAS = [
    # Manifiesto vacío
    _fila("", "2026-01-10", "AGENCIA NORTE", "BOGOTA(Cund)", "CUCUTA(Nort)",
          "STU901", "SENCILLO", "REM-9001", 1000000,
          "CONDUCTOR X", "10101010", "3000000001",
          "CLIENTE X", 200000, 0, "RESPONSABLE X"),
    # Manifiesto con texto (no numérico)
    _fila("ABC-ERROR", "2026-01-11", "AGENCIA SUR", "MEDELLIN(Anti)", "CALI(Vall)",
          "VWX234", "DOBLE TROQUE", "REM-9002", 1200000,
          "CONDUCTOR Y", "20202020", "3000000002",
          "CLIENTE Y", 300000, 50000, "RESPONSABLE Y"),
    # Fila completamente vacía (todas las celdas en blanco)
    _fila("", "", "", "", "", "", "", "", "",
          "", "", "", "", "", "", "", "", ""),
    # Manifiesto con valor especial (guión / N/A)
    _fila("N/A", "2026-01-12", "AGENCIA NORTE", "CALI(Vall)", "PEREIRA(Risa)",
          "YZA567", "SENCILLO", "REM-9003", 900000,
          "CONDUCTOR Z", "30303030", "3000000003",
          "CLIENTE Z", 250000, 0, "RESPONSABLE Z"),
]


def generar_invalidas(output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista_Manifiestos"

    _write_headers(ws)

    row = 2
    for fila in FILAS_VALIDAS:
        _write_row(ws, row, fila, invalida=False)
        row += 1

    for fila in FILAS_INVALIDAS:
        _write_row(ws, row, fila, invalida=True)
        row += 1

    _auto_width(ws)
    wb.save(output_path)
    return output_path


def generar_vacio(output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista_Manifiestos"

    _write_headers(ws)
    # Sin filas de datos — solo cabeceras

    _auto_width(ws)
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Genera Excels de prueba para el panel de carga del dashboard"
    )
    parser.add_argument("--invalidas", action="store_true",
                        help="Genera Excel con 6 filas válidas + 4 inválidas")
    parser.add_argument("--vacio", action="store_true",
                        help="Genera Excel vacío (solo cabeceras)")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Directorio de salida (default: tests/reportes/)")
    args = parser.parse_args()

    if not args.invalidas and not args.vacio:
        parser.print_help()
        print("\nEspecifica al menos --invalidas o --vacio (o ambos).")
        return

    # data/data_test/ en la raíz del proyecto
    proyecto_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    reportes_dir = args.output_dir or os.path.join(proyecto_root, "data", "data_test")
    os.makedirs(reportes_dir, exist_ok=True)

    if args.invalidas:
        path = os.path.join(reportes_dir, "prueba_invalidas.xlsx")
        generar_invalidas(path)
        print(f"Generado: {path}")
        print(f"  {len(FILAS_VALIDAS)} filas validas (manifiestos 999001–999006)")
        print(f"  {len(FILAS_INVALIDAS)} filas invalidas (en rojo en el Excel)")
        print("  Comportamiento esperado en el dashboard:")
        print("    - 6 manifiestos clasificados como 'Nuevos'")
        print("    - 4 filas reportadas como error/ignoradas sin trabar la carga")

    if args.vacio:
        path = os.path.join(reportes_dir, "prueba_vacio.xlsx")
        generar_vacio(path)
        print(f"Generado: {path}")
        print("  0 filas de datos (solo cabeceras)")
        print("  Comportamiento esperado en el dashboard:")
        print("    - Mensaje de error claro: 'El archivo no contiene datos'")


if __name__ == "__main__":
    main()
